import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('data/above-ground-gold-stocks.xlsx')
sheet = wb.active

# Extract years from row 3 (columns B to P)
years = [cell for cell in list(sheet.iter_rows(min_row=3, max_row=3, min_col=2, max_col=16, values_only=True))[0]]

# Extract data for each category
jewellery = list(sheet.iter_rows(min_row=4, max_row=4, min_col=2, max_col=16, values_only=True))[0]
central_banks = list(sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=16, values_only=True))[0]
private_investment = list(sheet.iter_rows(min_row=6, max_row=6, min_col=2, max_col=16, values_only=True))[0]
bars_coins = list(sheet.iter_rows(min_row=7, max_row=7, min_col=2, max_col=16, values_only=True))[0]
etfs = list(sheet.iter_rows(min_row=8, max_row=8, min_col=2, max_col=16, values_only=True))[0]
other = list(sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=16, values_only=True))[0]
total = list(sheet.iter_rows(min_row=10, max_row=10, min_col=2, max_col=16, values_only=True))[0]

# Build the timeline data structure
timeline_data = []

for i, year in enumerate(years):
    total_gold = total[i]
    jewellery_tonnes = jewellery[i]
    central_banks_tonnes = central_banks[i]
    bars_coins_tonnes = bars_coins[i]
    other_tonnes = other[i]
    
    # Calculate percentages
    jewellery_pct = (jewellery_tonnes / total_gold) * 100
    central_banks_pct = (central_banks_tonnes / total_gold) * 100
    bars_coins_pct = (bars_coins_tonnes / total_gold) * 100
    other_pct = (other_tonnes / total_gold) * 100
    
    # Calculate cube size (assuming 1m³ = 19,300 tonnes for gold density)
    cube_size = (total_gold / 19300) ** (1/3)
    
    year_data = {
        "year": int(year),
        "totalGold": round(total_gold, 0),
        "cubeSize": round(cube_size, 2),
        "categories": [
            {
                "name": "Jewellery",
                "tonnes": round(jewellery_tonnes, 0),
                "percentage": round(jewellery_pct, 1),
                "color": "#FFD700",
                "description": "Gold used in jewelry and ornaments"
            },
            {
                "name": "Bars & Coins",
                "tonnes": round(bars_coins_tonnes, 0),
                "percentage": round(bars_coins_pct, 1),
                "color": "#FFA500",
                "description": "Physical investment gold in bar and coin form"
            },
            {
                "name": "Central Banks",
                "tonnes": round(central_banks_tonnes, 0),
                "percentage": round(central_banks_pct, 1),
                "color": "#DAA520",
                "description": "Official gold reserves held by central banks"
            },
            {
                "name": "Other",
                "tonnes": round(other_tonnes, 0),
                "percentage": round(other_pct, 1),
                "color": "#B8860B",
                "description": "Electronics, dentistry, and other industrial uses"
            }
        ]
    }
    
    timeline_data.append(year_data)

# Print as JavaScript code
print("// ACTUAL DATA from World Gold Council (2010-2024)")
print("// Source: above-ground-gold-stocks.xlsx")
print("// Metals Focus, Refinitiv GFMS, World Gold Council")
print("window.goldDistributionTimeline = " + json.dumps(timeline_data, indent=2) + ";")
print("\n// Set current data to most recent year")
print("window.goldDistributionData = window.goldDistributionTimeline[window.goldDistributionTimeline.length - 1];")
